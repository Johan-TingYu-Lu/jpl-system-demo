import re, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

MD_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'full-detail-19.md')
OUT_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'full-detail-19.xlsx')

with open(MD_PATH, 'r', encoding='utf-8') as f:
    content = f.read()

# Parse student sections
students = re.split(r'\n---\n## (\d+ .+)', content)[1:]
student_blocks = [(students[i], students[i+1]) for i in range(0, len(students), 2)]

wb = Workbook()
wb.remove(wb.active)

header_font = Font(name='Arial', bold=True, size=11)
header_fill = PatternFill('solid', fgColor='D9E1F2')
warn_fill = PatternFill('solid', fgColor='FFF2CC')
err_fill = PatternFill('solid', fgColor='FCE4EC')
ok_fill = PatternFill('solid', fgColor='E8F5E9')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def parse_table(text):
    lines = [l.strip() for l in text.strip().split('\n') if l.strip().startswith('|')]
    if len(lines) < 2:
        return [], []
    headers = [c.strip() for c in lines[0].split('|')[1:-1]]
    rows = []
    for line in lines[2:]:
        cells = [c.strip() for c in line.split('|')[1:-1]]
        rows.append(cells)
    return headers, rows

def parse_code_block(text):
    m = re.search(r'```\n(.*?)```', text, re.DOTALL)
    if not m:
        return []
    lines = m.group(1).strip().split('\n')
    result = []
    for line in lines:
        if ':' in line:
            key, val = line.split(':', 1)
            result.append((key.strip(), val.strip()))
    return result

def write_headers(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def write_row(ws, row, values):
    for c, v in enumerate(values, 1):
        try:
            if v and v.startswith('$'):
                v = float(v.replace('$', '').replace(',', ''))
            elif v and v.replace('.','').replace('-','').isdigit():
                v = float(v) if '.' in v else int(v)
        except (ValueError, AttributeError):
            pass
        cell = ws.cell(row=row, column=c, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True)

# --- Sheet 1: Summary ---
ws_sum = wb.create_sheet('總覽')
sum_headers = ['#', 'ID', '姓名', '班別', '應製單數', 'Sheet計費', 'DB數量', '一致', 'Sheet已繳', 'DB已繳']
write_headers(ws_sum, 1, sum_headers)
for idx, (title, block) in enumerate(student_blocks):
    parts = title.strip().split(' ', 1)
    sid = parts[0]
    rest = parts[1] if len(parts) > 1 else ''
    name_match = re.match(r'(.+?)\s*\((.+?)\)', rest)
    name = name_match.group(1) if name_match else rest
    cls = name_match.group(2) if name_match else ''

    # Extract summary data
    summary_match = re.search(r'應製單數\(P\):\s*(\S+)', block)
    should_make = summary_match.group(1) if summary_match else ''
    bil_match = re.search(r'Sheet 計費日期表 \(應發單次數: (\d+)\)', block)
    sheet_count = int(bil_match.group(1)) if bil_match else 0
    db_match = re.search(r'DB Invoices \((\d+) 筆\)', block)
    db_count = int(db_match.group(1)) if db_match else 0
    match = '✓' if sheet_count == db_count else '✗'

    pay_match = re.search(r'Sheet 繳費日期表.*?\n\n(.*?)(?=\n###|\n---|\Z)', block, re.DOTALL)
    sheet_paid = 0
    if pay_match:
        for line in pay_match.group(1).split('\n'):
            if '|' in line and '未繳' not in line and '繳費日(raw)' not in line and '---' not in line:
                cells = [c.strip() for c in line.split('|')[1:-1]]
                if len(cells) >= 3 and cells[2] and cells[2] != '未繳':
                    sheet_paid += 1

    db_paid = 0
    inv_section = re.search(r'### DB Invoices.*?\n\n(.*?)(?=\n###|\n---|\Z)', block, re.DOTALL)
    if inv_section:
        for line in inv_section.group(1).split('\n'):
            if '| paid |' in line:
                db_paid += 1

    row_data = [idx+1, sid, name, cls, should_make, sheet_count, db_count, match, sheet_paid, db_paid]
    write_row(ws_sum, idx+2, row_data)
    if match == '✗':
        for c in range(1, len(row_data)+1):
            ws_sum.cell(row=idx+2, column=c).fill = err_fill

ws_sum.column_dimensions['A'].width = 4
ws_sum.column_dimensions['B'].width = 6
ws_sum.column_dimensions['C'].width = 12
ws_sum.column_dimensions['D'].width = 16
ws_sum.column_dimensions['E'].width = 10
ws_sum.column_dimensions['F'].width = 10
ws_sum.column_dimensions['G'].width = 8
ws_sum.column_dimensions['H'].width = 6
ws_sum.column_dimensions['I'].width = 10
ws_sum.column_dimensions['J'].width = 8

# --- Per-student sheets ---
for title, block in student_blocks:
    parts = title.strip().split(' ', 1)
    sid = parts[0]
    rest = parts[1] if len(parts) > 1 else ''
    name_match = re.match(r'(.+?)\s*\(', rest)
    name = name_match.group(1) if name_match else rest[:6]
    sheet_name = f'{sid} {name}'[:31]
    ws = wb.create_sheet(sheet_name)
    row = 1

    # Section: 學費收支總表
    ws.cell(row=row, column=1, value='學費收支總表').font = Font(name='Arial', bold=True, size=13)
    row += 1
    kv_pairs = parse_code_block(block)
    for key, val in kv_pairs:
        ws.cell(row=row, column=1, value=key).font = Font(name='Arial', bold=True)
        ws.cell(row=row, column=2, value=val)
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2).border = thin_border
        if key == '應製單數(P)' and val != '0':
            ws.cell(row=row, column=2).fill = warn_fill
        row += 1
    row += 1

    # Section: 計費日期表
    bil_section = re.search(r'### Sheet 計費日期表.*?\n\n(.*?)(?=\n###|\Z)', block, re.DOTALL)
    ws.cell(row=row, column=1, value='Sheet 計費日期表').font = Font(name='Arial', bold=True, size=13)
    row += 1
    if bil_section and '|' in bil_section.group(1):
        headers, rows = parse_table(bil_section.group(1))
        write_headers(ws, row, headers)
        row += 1
        for r in rows:
            write_row(ws, row, r)
            row += 1
    row += 1

    # Section: 繳費金額表
    fee_section = re.search(r'### Sheet 繳費金額表.*?\n\n(.*?)(?=\n###|\Z)', block, re.DOTALL)
    ws.cell(row=row, column=1, value='Sheet 繳費金額表').font = Font(name='Arial', bold=True, size=13)
    row += 1
    if fee_section and '|' in fee_section.group(1):
        headers, rows = parse_table(fee_section.group(1))
        write_headers(ws, row, headers)
        row += 1
        for r in rows:
            write_row(ws, row, r)
            row += 1
    row += 1

    # Section: 繳費日期表
    pay_section = re.search(r'### Sheet 繳費日期表.*?\n\n(.*?)(?=\n###|\Z)', block, re.DOTALL)
    ws.cell(row=row, column=1, value='Sheet 繳費日期表').font = Font(name='Arial', bold=True, size=13)
    row += 1
    if pay_section and '|' in pay_section.group(1):
        headers, rows = parse_table(pay_section.group(1))
        write_headers(ws, row, headers)
        row += 1
        for r in rows:
            write_row(ws, row, r)
            row += 1
    row += 1

    # Section: DB Invoices
    inv_section = re.search(r'### DB Invoices.*?\n\n(.*?)(?=\n###|\Z)', block, re.DOTALL)
    ws.cell(row=row, column=1, value='DB Invoices').font = Font(name='Arial', bold=True, size=13)
    row += 1
    if inv_section and '|' in inv_section.group(1):
        headers, rows = parse_table(inv_section.group(1))
        write_headers(ws, row, headers)
        row += 1
        for r in rows:
            write_row(ws, row, r)
            row += 1
    row += 1

    # Section: DB Payments
    pay_db_section = re.search(r'### DB Payments.*?\n\n(.*?)(?=\n###|\Z)', block, re.DOTALL)
    ws.cell(row=row, column=1, value='DB Payments').font = Font(name='Arial', bold=True, size=13)
    row += 1
    if pay_db_section and '|' in pay_db_section.group(1):
        headers, rows = parse_table(pay_db_section.group(1))
        write_headers(ws, row, headers)
        row += 1
        for r in rows:
            write_row(ws, row, r)
            row += 1
    row += 1

    # Section: DB 出席紀錄
    att_section = re.search(r'### DB 出席紀錄.*?\n\n(.*?)(?=\n###|\Z)', block, re.DOTALL)
    ws.cell(row=row, column=1, value='DB 出席紀錄').font = Font(name='Arial', bold=True, size=13)
    row += 1
    if att_section and '|' in att_section.group(1):
        headers, rows = parse_table(att_section.group(1))
        write_headers(ws, row, headers)
        row += 1
        for r in rows:
            write_row(ws, row, r)
            row += 1
    row += 1

    # Section: Invoice 計費明細
    detail_parts = re.findall(r'\*\*(.+?)\*\*.*?:\n\n(\|.*?)(?=\n\*\*|\n---|\n##|\Z)', block, re.DOTALL)
    if detail_parts:
        ws.cell(row=row, column=1, value='Invoice 計費明細').font = Font(name='Arial', bold=True, size=13)
        row += 1
        for label, table_text in detail_parts:
            ws.cell(row=row, column=1, value=label).font = Font(name='Arial', bold=True, size=10)
            row += 1
            headers, rows_data = parse_table(table_text)
            if headers:
                write_headers(ws, row, headers)
                row += 1
                for r in rows_data:
                    write_row(ws, row, r)
                    row += 1
            row += 1

    # Auto-width for key columns
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

wb.save(OUT_PATH)
print(f'Saved to {OUT_PATH} ({len(student_blocks)} student sheets + summary)')
